# report_dashboard.py

import streamlit as st
import pandas as pd
import altair as alt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from zipfile import ZipFile
import os

def report_dashboard():
    st.title("📊 Monthly Report Dashboard")
    uploaded_file = st.file_uploader("Upload Monthly Excel Report", type=[".xlsx"])

    if uploaded_file:
        try:
            excel_file = pd.ExcelFile(uploaded_file)
            sheet_names = excel_file.sheet_names
            selected_sheet = st.selectbox("Select Sheet", sheet_names)

            wb = load_workbook(uploaded_file, data_only=True)
            ws = wb[selected_sheet]

            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            visible_col_info = [
                (get_column_letter(cell.column), cell.value)
                for cell in header_row
                if cell.value is not None and not ws.column_dimensions[get_column_letter(cell.column)].hidden
            ]
            visible_letters = [col[0] for col in visible_col_info]
            visible_headers = [col[1] for col in visible_col_info]

            visible_data = []
            for row in ws.iter_rows(min_row=2):
                if not ws.row_dimensions[row[0].row].hidden:
                    row_data = [cell.value for cell in row if get_column_letter(cell.column) in visible_letters]
                    if any(cell is not None and cell != "" for cell in row_data):
                        visible_data.append(row_data)

            df = pd.DataFrame(visible_data, columns=visible_headers)

            # Clean Date Column
            if 'Date' in df.columns:
                df['Date'] = pd.to_datetime(df['Date'], errors='coerce', dayfirst=True)
                df = df.dropna(subset=['Date'])

            if 'Department' in df.columns:
                departments = df['Department'].dropna().unique().tolist()
                selected_dept = st.selectbox("Filter by Department", departments)
                df = df[df['Department'] == selected_dept]

            st.dataframe(df, use_container_width=True)

            # Download CSV
            csv = df.to_csv(index=False).encode('utf-8')
            st.download_button("⬇️ Download Filtered CSV", csv, f"{selected_sheet}_filtered.csv", "text/csv")

            # KPI Summary
            st.subheader("📌 KPI Summary")
            if 'Amount' in df.columns:
                st.metric("Total Amount", f"{df['Amount'].sum():,.2f}")
            if 'Activity' in df.columns:
                st.metric("Unique Activities", df['Activity'].nunique())

            # Charts
            numeric_cols = df.select_dtypes(include=['number']).columns.tolist()
            non_numeric_cols = df.select_dtypes(exclude=['number']).columns.tolist()

            if numeric_cols and non_numeric_cols:
                x_col = st.selectbox("X-axis (label/category)", non_numeric_cols)
                y_col = st.selectbox("Y-axis (value)", numeric_cols)
                chart_type = st.selectbox("Chart Type", ["Bar", "Line", "Pie"])

                chart_data = df[[x_col, y_col]].dropna()

                if chart_type == "Bar":
                    chart = alt.Chart(chart_data).mark_bar().encode(
                        x=alt.X(x_col, sort='-y'),
                        y=y_col,
                        color=x_col,
                        tooltip=[x_col, y_col]
                    )
                elif chart_type == "Line":
                    chart = alt.Chart(chart_data).mark_line(point=True).encode(
                        x=x_col,
                        y=y_col,
                        color=x_col,
                        tooltip=[x_col, y_col]
                    )
                elif chart_type == "Pie":
                    chart = alt.Chart(chart_data).mark_arc().encode(
                        theta=alt.Theta(field=y_col, type='quantitative'),
                        color=alt.Color(field=x_col, type='nominal'),
                        tooltip=[x_col, y_col]
                    )
                else:
                    st.warning("Unknown chart type")
                    return

                st.altair_chart(chart.properties(width=700, height=400))

            # Archive Summary
            if st.button("📦 Archive This Report"):
                archive_path = f"archived_reports/{selected_sheet}_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.csv"
                df.to_csv(archive_path, index=False)
                st.success(f"Archived to: {archive_path}")

        except Exception as e:
            st.error(f"Failed to process file: {e}")
