# Enhanced app.py with clean chart rendering and error-free structure

import streamlit as st
import pandas as pd
import altair as alt
from zipfile import ZipFile
from openpyxl import load_workbook

st.set_page_config(page_title="Monthly Report Dashboard", layout="wide")

st.markdown("""
    <h1 style='text-align: center; color: #1f77b4;'>📊 Monthly Departmental Report Dashboard</h1>
""", unsafe_allow_html=True)

uploaded_file = st.file_uploader("Upload the Excel report", type=[".xlsx"])

if uploaded_file:
    try:
        excel_file = pd.ExcelFile(uploaded_file)
        sheet_names = excel_file.sheet_names

        st.markdown("### 🗂 Select a Sheet to View")
        selected_sheet = st.selectbox("Choose a sheet", sheet_names)

        wb = load_workbook(uploaded_file, data_only=True)
        ws = wb[selected_sheet]

        # Collect visible columns and their letter indexes
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        visible_col_info = [(cell.column_letter, cell.value) for cell in header_row if cell.value is not None and not ws.column_dimensions[cell.column_letter].hidden]
        visible_letters = [col[0] for col in visible_col_info]
        visible_headers = [col[1] for col in visible_col_info]

        # Prepare consistent data from visible columns
        visible_data = []
        for row in ws.iter_rows(min_row=2):
            if not ws.row_dimensions[row[0].row].hidden:
                row_data = [cell.value for cell in row if cell.column_letter in visible_letters]
                if any(cell is not None and cell != "" for cell in row_data):
                    visible_data.append(row_data)

        df = pd.DataFrame(visible_data, columns=visible_headers)

        if 'Department' in df.columns:
            departments = df['Department'].dropna().unique().tolist()
            selected_dept = st.selectbox(f"Filter by Department in '{selected_sheet}'", departments)
            df = df[df['Department'] == selected_dept]

        st.dataframe(df, use_container_width=True)

        csv = df.to_csv(index=False).encode('utf-8')
        st.download_button(
            label="⬇️ Download Table as CSV",
            data=csv,
            file_name=f"{selected_sheet}_filtered_data.csv",
            mime="text/csv"
        )

        numeric_columns = df.select_dtypes(include=['number']).columns.tolist()
        non_numeric_columns = df.select_dtypes(exclude=['number']).columns.tolist()
        default_label = non_numeric_columns[0] if non_numeric_columns else None
        label_column = st.selectbox("Select label column (x-axis / category):", [None] + non_numeric_columns, index=(non_numeric_columns.index(default_label) + 1) if default_label else 0)
        value_column = st.selectbox("Select column for values (y-axis):", numeric_columns)

        if label_column:
            unique_labels = df[label_column].dropna().tolist()
            chart_data = df[df[label_column].isin(unique_labels)][[label_column, value_column]].dropna()
            chart_data[label_column] = pd.Categorical(chart_data[label_column], categories=unique_labels, ordered=True)
            chart_data = chart_data.sort_values(label_column)
        else:
            chart_data = df[[value_column]].dropna().reset_index()
            chart_data.rename(columns={'index': 'index_label'}, inplace=True)
            chart_data['index_label'] = chart_data['index_label'].astype(str)
            label_column = 'index_label'
            chart_data[label_column] = chart_data[label_column]

        st.markdown("### 📊 Selected Data Preview")
        preview_data = chart_data.copy()
        if preview_data[value_column].max() <= 1 and preview_data[value_column].min() >= 0:
            preview_data[value_column] = (preview_data[value_column] * 100).round(0).astype(int).astype(str) + '%'
        else:
            preview_data[value_column] = preview_data[value_column].apply(lambda x: f"{x:,.0f}")
        st.dataframe(preview_data[[label_column, value_column]], use_container_width=True)

        st.markdown("### 🎨 Chart Styling")
        color_scheme = st.selectbox("Choose a color theme:", ["category10", "category20", "tableau10", "accent", "dark2"], index=0)

        chart_types = st.multiselect(
            "Select chart types to display:",
            ["Bar Chart", "Line Chart", "Pie Chart"],
            default=["Bar Chart"]
        )

        chart_width = st.slider("Chart width", 400, 1000, 700)
        chart_height = st.slider("Chart height", 200, 600, 300)

        if not chart_data.empty:
            tooltip_vals = [label_column, alt.Tooltip(f"{value_column}:Q", title="Value", format=",.0f")]

            if "Bar Chart" in chart_types:
                st.markdown(f"#### 🔢 Bar Chart for: {value_column}")
                bar_chart = alt.Chart(chart_data).mark_bar().encode(
                    x=alt.X(f"{label_column}:O", sort=unique_labels),
                    y=alt.Y(f"{value_column}:Q"),
                    color=alt.Color(f"{label_column}:N", scale=alt.Scale(scheme=color_scheme)),
                    tooltip=tooltip_vals
                ).properties(width=chart_width, height=chart_height)
                st.altair_chart(bar_chart)

            if "Line Chart" in chart_types:
                st.markdown(f"#### 📈 Line Chart for: {value_column}")
                line_chart = alt.Chart(chart_data).mark_line(point=True).encode(
                    x=alt.X(f"{label_column}:O", sort=unique_labels),
                    y=alt.Y(f"{value_column}:Q"),
                    color=alt.Color(f"{label_column}:N", scale=alt.Scale(scheme=color_scheme)),
                    tooltip=tooltip_vals
                ).properties(width=chart_width, height=chart_height)
                st.altair_chart(line_chart)

            if "Pie Chart" in chart_types:
                st.markdown(f"#### 🥰 Pie Chart (Donut) for: {value_column}")

                chart_data['label_display'] = chart_data[label_column].astype(str)

                pie_chart = alt.Chart(chart_data).mark_arc(innerRadius=60).encode(
                    theta=alt.Theta(field=value_column, type='quantitative'),
                    color=alt.Color(field='label_display', type='nominal', scale=alt.Scale(scheme=color_scheme)),
                    tooltip=[label_column, alt.Tooltip(f'{value_column}:Q', title='Value', format=',.0f')]
                ).properties(width=chart_height, height=chart_height)

                st.altair_chart(pie_chart)
        else:
            st.info("ℹ️ No data selected for chart generation.")

        with ZipFile("workbook_export.zip", "w") as zipf:
            for sheet in sheet_names:
                ws = wb[sheet]
                header_row = next(ws.iter_rows(min_row=1, max_row=1))
                visible_col_info = [(cell.column_letter, cell.value) for cell in header_row if cell.value is not None and not ws.column_dimensions[cell.column_letter].hidden]
                visible_letters = [col[0] for col in visible_col_info]
                visible_headers = [col[1] for col in visible_col_info]

                visible_rows = []
                for row in ws.iter_rows(min_row=2):
                    if not ws.row_dimensions[row[0].row].hidden:
                        row_data = [cell.value for cell in row if cell.column_letter in visible_letters]
                        if any(cell is not None and cell != "" for cell in row_data):
                            visible_rows.append(row_data)

                df_sheet = pd.DataFrame(visible_rows, columns=visible_headers)
                csv_bytes = df_sheet.to_csv(index=False).encode("utf-8")
                zipf.writestr(f"{sheet}.csv", csv_bytes)

        with open("workbook_export.zip", "rb") as f:
            st.download_button(
                label="⬇️ Download Entire Workbook as ZIP of CSVs",
                data=f.read(),
                file_name="Workbook_Export.zip",
                mime="application/zip"
            )

    except Exception as e:
        st.error(f"❌ Failed to read Excel file: {e}")
else:
    st.warning("Please upload a valid Excel report to continue.")
